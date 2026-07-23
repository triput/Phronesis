#!/usr/bin/env python3
"""Generate Phronesis V3 automated test inventory (CSV + optional XLSX).

Scans phronesis_app/tests/**/*.py for unittest/Django TestCase methods named
test_* and emits:
  docs/V3_AUTOMATED_TEST_INVENTORY.csv
  docs/V3_AUTOMATED_TEST_INVENTORY.xlsx  (when openpyxl is installed)

Run from repo root:
  python tool/generate_test_inventory.py
"""

from __future__ import annotations

import csv
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TEST_DIR = ROOT / "phronesis_app" / "tests"
OUT_CSV = ROOT / "docs" / "V3_AUTOMATED_TEST_INVENTORY.csv"
OUT_XLSX = ROOT / "docs" / "V3_AUTOMATED_TEST_INVENTORY.xlsx"
TODAY = date.today().isoformat()

COLUMNS = [
    "test_id",
    "phase",
    "tier_refs",
    "test_file",
    "test_class",
    "test_name",
    "kind",
    "component",
    "platform",
    "asserts",
    "status",
    "evaluation_status",
    "added_date",
    "last_verified_phase",
    "manual_companion",
    "notes",
]

# File-level defaults. Individual rows inherit these unless heuristics override kind.
FILE_META: dict[str, dict[str, str]] = {
    "test_p0_foundation.py": {
        "phase": "P0",
        "tier_refs": "VN-A;Foundation",
        "kind": "smoke",
        "component": "Foundation",
        "platform": "Web",
    },
    "test_p0_auth.py": {
        "phase": "P0",
        "tier_refs": "VN-A;Security",
        "kind": "request",
        "component": "Auth / Security",
        "platform": "Web",
    },
    "test_p0_modules.py": {
        "phase": "P0",
        "tier_refs": "VN-A03;Modules",
        "kind": "request",
        "component": "Modules / Presets",
        "platform": "Web",
    },
    "test_p0_backup.py": {
        "phase": "P0",
        "tier_refs": "VN-A05;Backup;S-41",
        "kind": "request",
        "component": "Backup / Restore",
        "platform": "Web",
    },
    "test_p0_standalone_paths.py": {
        "phase": "P0",
        "tier_refs": "VN-B01;Standalone",
        "kind": "unit",
        "component": "Windows / Paths",
        "platform": "Web",
    },
    "test_p0_trash.py": {
        "phase": "P0",
        "tier_refs": "VN-A07;Trash",
        "kind": "request",
        "component": "Trash",
        "platform": "Web",
    },
    "test_p1_capture_focus.py": {
        "phase": "P1",
        "tier_refs": "VN-A04;Capture",
        "kind": "request",
        "component": "Capture / Focus",
        "platform": "Web",
    },
    "test_p2_matrix.py": {
        "phase": "P2",
        "tier_refs": "Matrix",
        "kind": "unit",
        "component": "Matrix / Drawer",
        "platform": "Web",
    },
    "test_p3_settings.py": {
        "phase": "P3",
        "tier_refs": "Settings",
        "kind": "request",
        "component": "Settings",
        "platform": "Web",
    },
    "test_p3_planning.py": {
        "phase": "P3",
        "tier_refs": "Today;Plan",
        "kind": "request",
        "component": "Planning / Today",
        "platform": "Web",
    },
    "test_p3_calendar.py": {
        "phase": "P3",
        "tier_refs": "Calendar",
        "kind": "unit",
        "component": "Calendar",
        "platform": "Web",
    },
    "test_p3_calendar_microsoft.py": {
        "phase": "P3",
        "tier_refs": "Calendar;OAuth",
        "kind": "integration",
        "component": "Calendar / Microsoft",
        "platform": "Web",
    },
    "test_p3_calendar_grid.py": {
        "phase": "P3",
        "tier_refs": "Calendar",
        "kind": "request",
        "component": "Calendar / Grid",
        "platform": "Web",
    },
    "test_p3_alerts.py": {
        "phase": "P3",
        "tier_refs": "Alerts",
        "kind": "request",
        "component": "Alerts",
        "platform": "Web",
    },
    "test_p3_time_locale.py": {
        "phase": "P3",
        "tier_refs": "Locale",
        "kind": "unit",
        "component": "Time / Locale",
        "platform": "Web",
    },
    "test_p3_telemetry.py": {
        "phase": "P3",
        "tier_refs": "Telemetry",
        "kind": "unit",
        "component": "Telemetry",
        "platform": "Web",
    },
    "test_p3_telemetry_bands.py": {
        "phase": "P3",
        "tier_refs": "Telemetry",
        "kind": "unit",
        "component": "Telemetry / Bands",
        "platform": "Web",
    },
    "test_p3_tele_geocode.py": {
        "phase": "P3",
        "tier_refs": "Telemetry",
        "kind": "unit",
        "component": "Telemetry / Geocode",
        "platform": "Web",
    },
    "test_p3_tele_location.py": {
        "phase": "P3",
        "tier_refs": "Telemetry",
        "kind": "unit",
        "component": "Telemetry / Location",
        "platform": "Web",
    },
    "test_p3_appearance.py": {
        "phase": "P3",
        "tier_refs": "Appearance",
        "kind": "request",
        "component": "Appearance",
        "platform": "Web",
    },
    "test_p4_due_pulse.py": {
        "phase": "P4",
        "tier_refs": "DuePulse",
        "kind": "unit",
        "component": "Due Pulse",
        "platform": "Web",
    },
    "test_p4_board.py": {
        "phase": "P4",
        "tier_refs": "Board",
        "kind": "request",
        "component": "Board",
        "platform": "Web",
    },
    "test_p4_academy.py": {
        "phase": "P4",
        "tier_refs": "Academy",
        "kind": "request",
        "component": "Academy",
        "platform": "Web",
    },
    "test_p4_overview.py": {
        "phase": "P4",
        "tier_refs": "Overview",
        "kind": "request",
        "component": "Overview",
        "platform": "Web",
    },
    "test_p4_views.py": {
        "phase": "P4",
        "tier_refs": "Home",
        "kind": "request",
        "component": "Views / Home",
        "platform": "Web",
    },
    "test_p4_stability.py": {
        "phase": "P4",
        "tier_refs": "Stability",
        "kind": "unit",
        "component": "Stability",
        "platform": "Web",
    },
    "test_p4_analytics.py": {
        "phase": "P4",
        "tier_refs": "Analytics",
        "kind": "request",
        "component": "Analytics",
        "platform": "Web",
    },
    "test_p5_templates.py": {
        "phase": "P5",
        "tier_refs": "Templates",
        "kind": "unit",
        "component": "Templates",
        "platform": "Web",
    },
    "test_p5_recurrence.py": {
        "phase": "P5",
        "tier_refs": "Recurrence",
        "kind": "unit",
        "component": "Recurrence",
        "platform": "Web",
    },
    "test_p5_bulk_create.py": {
        "phase": "P5",
        "tier_refs": "BulkCreate",
        "kind": "request",
        "component": "Bulk Create",
        "platform": "Web",
    },
    "test_p5_celery.py": {
        "phase": "P5",
        "tier_refs": "Scheduler",
        "kind": "integration",
        "component": "Celery / Scheduler",
        "platform": "Web",
    },
    "test_p5_reminders.py": {
        "phase": "P5",
        "tier_refs": "Reminders",
        "kind": "unit",
        "component": "Reminders",
        "platform": "Web",
    },
    "test_p5_calendar_push.py": {
        "phase": "P5",
        "tier_refs": "Calendar",
        "kind": "integration",
        "component": "Calendar / Push",
        "platform": "Web",
    },
    "test_p5_a11y.py": {
        "phase": "P5",
        "tier_refs": "A11y",
        "kind": "request",
        "component": "Accessibility",
        "platform": "Web",
    },
    "test_bl_rec_001.py": {
        "phase": "BL",
        "tier_refs": "BL-REC-001",
        "kind": "unit",
        "component": "Recurrence / Backlog",
        "platform": "Web",
    },
    "test_bl_rec_002.py": {
        "phase": "BL",
        "tier_refs": "BL-REC-002",
        "kind": "unit",
        "component": "Recurrence / Backlog",
        "platform": "Web",
    },
    "test_bl_time_004.py": {
        "phase": "BL",
        "tier_refs": "BL-TIME-004",
        "kind": "unit",
        "component": "Time / Backlog",
        "platform": "Web",
    },
    "test_notify_adapters.py": {
        "phase": "Notify",
        "tier_refs": "Notifications",
        "kind": "unit",
        "component": "Notifications",
        "platform": "Web",
    },
    "test_polish_optional.py": {
        "phase": "Polish",
        "tier_refs": "Polish",
        "kind": "request",
        "component": "Polish",
        "platform": "Web",
    },
}

PHASE_FROM_FILENAME: dict[str, str] = {
    "p0": "P0",
    "p1": "P1",
    "p2": "P2",
    "p3": "P3",
    "p4": "P4",
    "p5": "P5",
    "bl": "BL",
    "notify": "Notify",
    "polish": "Polish",
}

CLASS_RE = re.compile(r"^class\s+(\w+)\s*\([^)]*TestCase[^)]*\)\s*:", re.MULTILINE)
TEST_DEF_RE = re.compile(r"^(\s+)def (test_\w+)\s*\(", re.MULTILINE)
DOCSTRING_RE = re.compile(
    r'^\s+def test_\w+\s*\([^)]*\)\s*:\s*\n\s+("""|\'\'\')(.*?)\1',
    re.MULTILINE | re.DOTALL,
)


@dataclass
class Case:
    rel_path: str
    file_name: str
    test_class: str
    test_name: str
    line: int
    docstring: str
    uses_client: bool


def infer_phase(file_name: str) -> str:
    meta = FILE_META.get(file_name, {})
    if "phase" in meta:
        return meta["phase"]
    stem = file_name.replace("test_", "").split("_")[0].lower()
    return PHASE_FROM_FILENAME.get(stem, "Other")


def infer_kind(file_name: str, test_name: str, test_class: str, uses_client: bool) -> str:
    meta = FILE_META.get(file_name, {})
    default = meta.get("kind", "unit")
    name_lower = test_name.lower()
    class_lower = test_class.lower()
    if uses_client or "client" in class_lower or "view" in name_lower:
        return "request"
    if "integration" in name_lower or "celery" in name_lower or "sync" in name_lower:
        return "integration"
    if "smoke" in name_lower or default == "smoke":
        return "smoke"
    return default


def infer_asserts(docstring: str, test_name: str) -> str:
    if docstring:
        first = docstring.strip().splitlines()[0].strip()
        if first:
            return first
    # Human-readable fallback from snake_case name
    label = test_name.removeprefix("test_").replace("_", " ")
    return label[0].upper() + label[1:] if label else "—"


def manual_companion(phase: str, component: str, file_name: str) -> str:
    if phase == "P0" and "Auth" in component:
        return "V3_SECURITY_TEST_CHECKLIST"
    if phase == "P3" and "Calendar" in component:
        return "V3_MANUAL_E2E_MATRIX#Calendar"
    if phase == "P5" and file_name in {"test_p5_celery.py", "test_p5_reminders.py"}:
        return "V3_PERFORMANCE_TEST_CHECKLIST"
    if phase == "P4" and component.startswith("Views"):
        return "V3_PERFORMANCE_TEST_CHECKLIST"
    return ""


def parse_file(path: Path) -> list[Case]:
    text = path.read_text(encoding="utf-8")
    rel_path = path.relative_to(ROOT).as_posix()
    file_name = path.name

    # Map test method start positions to classes
    class_spans: list[tuple[int, str]] = []
    for m in CLASS_RE.finditer(text):
        class_spans.append((m.start(), m.group(1)))
    class_spans.sort(key=lambda x: x[0])

    def class_at(pos: int) -> str:
        current = "—"
        for start, name in class_spans:
            if start <= pos:
                current = name
            else:
                break
        return current

    # Docstrings keyed by test name (first match per method in file)
    doc_by_test: dict[str, str] = {}
    for m in TEST_DEF_RE.finditer(text):
        test_name = m.group(2)
        if test_name in doc_by_test:
            continue
        tail = text[m.end() : m.end() + 400]
        dm = re.match(r'\s*:\s*\n\s+("""|\'\'\')(.*?)\1', tail, re.DOTALL)
        if dm:
            doc_by_test[test_name] = dm.group(2).strip()

    cases: list[Case] = []
    for m in TEST_DEF_RE.finditer(text):
        test_name = m.group(2)
        line = text.count("\n", 0, m.start()) + 1
        test_class = class_at(m.start())
        # Heuristic: method body uses self.client within next ~80 lines
        body_start = m.end()
        body_end = text.find("\n    def ", body_start)
        if body_end == -1:
            body_end = min(len(text), body_start + 4000)
        body = text[body_start:body_end]
        uses_client = "self.client" in body or "Client()" in body
        cases.append(
            Case(
                rel_path=rel_path,
                file_name=file_name,
                test_class=test_class,
                test_name=test_name,
                line=line,
                docstring=doc_by_test.get(test_name, ""),
                uses_client=uses_client,
            )
        )
    return cases


def phase_sort_key(phase: str) -> tuple[int, str]:
    order = {
        "P0": 0,
        "P1": 1,
        "P2": 2,
        "P3": 3,
        "P4": 4,
        "P5": 5,
        "BL": 6,
        "Notify": 7,
        "Polish": 8,
        "Other": 9,
    }
    return (order.get(phase, 99), phase)


def build_rows() -> list[dict[str, str]]:
    all_cases: list[Case] = []
    for path in sorted(TEST_DIR.rglob("test_*.py")):
        all_cases.extend(parse_file(path))

    # Stable sort for deterministic IDs
    all_cases.sort(
        key=lambda c: (
            phase_sort_key(infer_phase(c.file_name)),
            c.rel_path,
            c.test_class,
            c.test_name,
        )
    )

    counters: Counter[str] = Counter()
    rows: list[dict[str, str]] = []
    for case in all_cases:
        meta = FILE_META.get(case.file_name, {})
        phase = infer_phase(case.file_name)
        counters[phase] += 1
        test_id = f"AT-{phase}-{counters[phase]:03d}"
        component = meta.get("component", "Unmapped")
        kind = infer_kind(case.file_name, case.test_name, case.test_class, case.uses_client)
        rows.append(
            {
                "test_id": test_id,
                "phase": phase,
                "tier_refs": meta.get("tier_refs", ""),
                "test_file": case.rel_path,
                "test_class": case.test_class,
                "test_name": case.test_name,
                "kind": kind,
                "component": component,
                "platform": meta.get("platform", "Web"),
                "asserts": infer_asserts(case.docstring, case.test_name),
                "status": "active",
                "evaluation_status": "Cataloged",
                "added_date": TODAY,
                "last_verified_phase": "",
                "manual_companion": manual_companion(phase, component, case.file_name),
                "notes": f"line {case.line}",
            }
        )
    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def style_header(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), max_width)


def write_xlsx(rows: list[dict[str, str]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Automated"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row[c] for c in COLUMNS])
    style_header(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    ws.freeze_panes = "A2"
    table = Table(
        displayName="AutomatedTests",
        ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}",
    )
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    autosize(ws)

    by_file = Counter(r["test_file"] for r in rows)
    wf = wb.create_sheet("By_File")
    wf.append(["test_file", "case_count", "phase", "kind", "component", "platform"])
    file_meta: dict[str, dict[str, str]] = {}
    for r in rows:
        file_meta.setdefault(
            r["test_file"],
            {
                "phase": r["phase"],
                "kind": r["kind"],
                "component": r["component"],
                "platform": r["platform"],
            },
        )
    for path, count in sorted(by_file.items()):
        meta = file_meta[path]
        wf.append([path, count, meta["phase"], meta["kind"], meta["component"], meta["platform"]])
    style_header(wf)
    wf.auto_filter.ref = f"A1:F{len(by_file) + 1}"
    wf.freeze_panes = "A2"
    autosize(wf)

    ws_sum = wb.create_sheet("Phase_Summary")
    ws_sum.append(["phase", "case_count", "file_count", "unit", "request", "integration", "smoke"])
    by_phase_cases = Counter(r["phase"] for r in rows)
    by_phase_files: dict[str, set[str]] = defaultdict(set)
    by_phase_kind: dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        by_phase_files[r["phase"]].add(r["test_file"])
        by_phase_kind[r["phase"]][r["kind"]] += 1
    for phase, count in sorted(by_phase_cases.items(), key=lambda kv: phase_sort_key(kv[0])):
        kinds = by_phase_kind[phase]
        ws_sum.append(
            [
                phase,
                count,
                len(by_phase_files[phase]),
                kinds.get("unit", 0),
                kinds.get("request", 0),
                kinds.get("integration", 0),
                kinds.get("smoke", 0),
            ]
        )
    style_header(ws_sum)
    autosize(ws_sum)

    wr = wb.create_sheet("Readme")
    wr.append(["Field", "Value"])
    info = [
        ("Canonical git artifact", "docs/V3_AUTOMATED_TEST_INVENTORY.csv"),
        ("Operator workbook", "docs/V3_AUTOMATED_TEST_INVENTORY.xlsx"),
        ("Generator", "tool/generate_test_inventory.py"),
        ("Generated", TODAY),
        ("Total cases", str(len(rows))),
        ("Total files", str(len(by_file))),
        ("Run suite", "python manage.py test phronesis_app.tests -v 1"),
        ("Update ritual", "Renee delta → Page regenerate → Steve gates phase land"),
        ("evaluation_status", "Cataloged = listed; Pass/Fail only after manage.py test run"),
        ("Manual companion", "docs/V3_MANUAL_E2E_MATRIX.csv"),
    ]
    for row in info:
        wr.append(list(row))
    style_header(wr)
    autosize(wr, max_width=80)

    wb.save(OUT_XLSX)
    return True


def main() -> int:
    rows = build_rows()
    write_csv(rows)
    xlsx_ok = write_xlsx(rows)
    by_phase = Counter(r["phase"] for r in rows)
    by_file = Counter(r["test_file"] for r in rows)
    print(f"Wrote {OUT_CSV}")
    if xlsx_ok:
        print(f"Wrote {OUT_XLSX}")
    else:
        print("Skipped XLSX (install openpyxl to enable workbook export)")
    print(f"cases={len(rows)} files={len(by_file)}")
    for phase, count in sorted(by_phase.items(), key=lambda kv: phase_sort_key(kv[0])):
        print(f"  {phase}: {count}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
